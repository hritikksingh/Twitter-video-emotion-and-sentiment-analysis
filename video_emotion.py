import numpy as np
import cv2
from fastai import *
from fastai.vision import *
import pandas as pd
from imutils.video import FileVideoStream
import time
import dlib
import openpyxl
import glob


#Function to run the video and write emotion on excel sheet
def run_video():
    #Loading Model files
    path = "./data/video/"
    vidcap = FileVideoStream(glob.glob('./video/*')[0]).start()
    learn = load_learner(path, 'export.pkl')
    face_cascade = cv2.CascadeClassifier("./data/video/haarcascade_frontalface_default.xml")
    predictor = dlib.shape_predictor("./data/video/shape_predictor_68_face_landmarks.dat")
    count = 0
    framecount = 0
    data=[]
    while vidcap.more():
        frame = vidcap.read()
        if frame is None:
            break
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        face_coord = face_cascade.detectMultiScale(gray, 1.1, 20, minSize=(30, 30))
        for coords in face_coord:
            X, Y, w, h = coords
            H, W, _ = frame.shape
            X_1, X_2 = (max(0, X - int(w * 0.3)), min(X + int(1.3 * w), W))
            Y_1, Y_2 = (max(0, Y - int(0.3 * h)), min(Y + int(1.3 * h), H))
            img_cp = gray[Y_1:Y_2, X_1:X_2].copy()
            if framecount % 10 == 0:
                prediction, idx, probability = learn.predict(Image(pil2tensor(img_cp, np.float32).div_(225)))
                data.append([prediction])
        framecount += 1
    df = pd.DataFrame(data, columns = [ 'Expression'])
    df.to_excel(path+'exportvideo.xlsx')
    vidcap.stop()
    
#function to calculate the percentage of emotions
def get_perc():
    wrkbk=openpyxl.load_workbook("./data/video/exportvideo.xlsx")
    sh=wrkbk.active
    count={}
    for i in range(2,sh.max_row+1):
        cell_val=sh.cell(row=i,column=2).value
        if cell_val not in count:
            count[cell_val]=1
        else:
            count[cell_val]+=1
    total=sh.max_row-1;
    for key,value in count.items():
        count[key]=(value/total)*100
    return count

#main fucntion to be called
def get_video_emotion():
    run_video()
    return get_perc()

